import tiktoken
import pandas as pd
import asyncio
import aiohttp
import nltk
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.core.files.temp import NamedTemporaryFile
from langchain.llms import OpenAI
import os
from time import sleep


nltk.download("stopwords")
from nltk.corpus import stopwords

stopwords_en = stopwords.words("english")


class BasePipeline:
    def count_tokens(self, text):
        encoding = tiktoken.get_encoding("cl100k_base")
        num_tokens = len(encoding.encode(text))
        return num_tokens

    async def async_gpt_completion_call(self, session, prompt, retry_count=0):
        sleep(10 * (2 ** retry_count) * min(retry_count, 1))
        llm = OpenAI(
            model_name="gpt-3.5-turbo",
            openai_api_key=os.environ.get("OPENAI_API_KEY"),
        )

        try:
            resp = await llm.agenerate([prompt])
            return resp.generations[0][0].text
        except Exception as e:
            print(e)
            if retry_count < 5:
                return await self.async_gpt_completion_call(session, prompt, retry_count+1)
            return {"error": e}
    
    def parse_comments(self, comments):
        all_str = ""
        for i, text in enumerate(comments):
            all_str += str(i + 1) + " " + text + "\n"
        return all_str

    async def get_top_down_topics(self, session, parsed_comments):
        prompt = self.top_down_topics_prompt + "\n" + parsed_comments
        data = await self.async_gpt_completion_call(session, prompt)
        if "error" in data:
            raise Exception(data["error"])

        top_down_topics = data

        topics_without_header_without_metadata = []
        for topic_text in top_down_topics.split("\n"):
            try:
                topics_without_header_without_metadata.append(
                    topic_text.split("(")[0]
                    .strip()
                    .split("-")[0]
                    .strip()
                    .split(".")[1]
                    .strip()
                )
            except:
                pass
        top_down_topics = topics_without_header_without_metadata
        return top_down_topics

    async def get_top_down_topics_tagging(
        self, session, parsed_comments, top_down_topics
    ):
        prompt = (
            self.top_down_topic_tagging_prompt
            + "\n"
            + "Theme dictionary: "
            + "\n"
            + "\n".join(top_down_topics)
            + "\n"
            + "YouTube comment list:"
            + "\n"
            + parsed_comments
        )
        data = await self.async_gpt_completion_call(session, prompt)
        if "error" in data:
            raise Exception(data["error"])

        completion = data
        return completion

    def adjust_token_limit(self, start_idx, end_idx, max_tokens, dec_amount):
        parsed_comments = self.parse_comments(self.all_comments[start_idx:end_idx])

        num_tokens = self.count_tokens(parsed_comments)
        while num_tokens > max_tokens:
            print("Decreasing end idx from: ", str(end_idx))
            end_idx = end_idx - dec_amount
            print("New end idx: ", str(end_idx))
            time.sleep(2)

            parsed_comments = self.parse_comments(self.all_comments[start_idx:end_idx])
            num_tokens = self.count_tokens(parsed_comments)

        return {"parsed_comments": parsed_comments, "end_idx": end_idx}

    async def get_analyses(self):
        start_idx = 0
        end_idx = min(start_idx + self.max_top_down_comments, len(self.all_comments))

        self.analysis_df = {
            "comments": [],
            "Top Down Topics": [],
            "Top Down Topics Tagged": [],
        }

        parsed_comments_list = []

        once_equal_break = 0
        while end_idx <= len(self.all_comments):
            if once_equal_break:
                break

            if end_idx == len(self.all_comments):
                once_equal_break = 1

            print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
            print("Top Down topics...")

            _temp = self.adjust_token_limit(
                start_idx, end_idx, self.max_top_down_length, 10
            )
            parsed_comments = _temp["parsed_comments"]
            end_idx = _temp["end_idx"]

            self.analysis_df["comments"].append(parsed_comments)
            print(
                "Within limit from start_idx: ",
                str(start_idx) + " to end_idx: ",
                str(end_idx),
            )

            # Fetch topics from start_idx to end_idx
            parsed_comments_list.append((start_idx, end_idx, parsed_comments))

            start_idx = end_idx
            end_idx = min(
                start_idx + self.max_top_down_comments, len(self.all_comments)
            )

        async with aiohttp.ClientSession() as session:
            async_responses = [
                self.top_down_topics_tagging(session, pair[0], pair[1], pair[2])
                for pair in parsed_comments_list
            ]
            all_topics = await asyncio.gather(*async_responses)

        # store topics
        self.analysis_df["Top Down Topics"].extend(all_topics)

        return self.analysis_df

    async def top_down_topics_tagging(
        self, session, start_idx, end_idx, parsed_comments
    ):
        print("----------------------------")
        print("Top Down and bottom up topic tagging...")

        mini_end_idx = min(start_idx + self.max_bottom_up_comments, end_idx)

        comments_list = []

        top_down_topics = await self.get_top_down_topics(session, parsed_comments)
        self.analysis_df["Top Down Topics"].extend(top_down_topics)

        # Topic tagging
        mini_once_equal_break = 0
        while mini_end_idx <= end_idx:
            if mini_once_equal_break:
                break

            if mini_end_idx == end_idx:
                mini_once_equal_break = 1

            print(
                "Topics tagging from start idx: ",
                str(start_idx),
                " to end idx: ",
                str(mini_end_idx),
            )
            _temp = self.adjust_token_limit(
                start_idx, mini_end_idx, self.max_bottom_up_length, 2
            )
            parsed_comments = _temp["parsed_comments"]
            mini_end_idx = _temp["end_idx"]

            print(
                "Within limit from start_idx: ",
                str(start_idx) + " to mini end_idx: ",
                str(mini_end_idx),
            )

            comments_list.append(parsed_comments)

            start_idx = mini_end_idx
            mini_end_idx = min(start_idx + self.max_bottom_up_comments, end_idx)

        async_responses = [
            self.get_top_down_topics_tagging(session, comment, top_down_topics)
            for comment in comments_list
        ]
        all_top_down_tagging = await asyncio.gather(*async_responses)

        all_top_down_tagging = "\n".join(all_top_down_tagging)
        self.analysis_df["Top Down Topics Tagged"].append(all_top_down_tagging)

    def parse_analyses(self):
        print("Parsing top down tagging")
        split_tags_td = []

        positive = 0
        negative = 0
        neutral = 0
        comment_sentiment_dict = []

        workbook = Workbook()
        del workbook["Sheet"]
        ws_stats = workbook.create_sheet("All Stats")
        ws_topics_match = workbook.create_sheet("Topics Match")
        ws_comment_sentiment = workbook.create_sheet("Top Down Mapping Full Match")

        for tag_m in self.analysis_df["Top Down Topics Tagged"]:
            for tag in tag_m.split("\n"):
                if "youtube comment" in tag.lower() or (
                    "mapped themes" in tag.lower() and "comment text" in tag.lower()
                ):
                    continue
                if "---" in tag:
                    continue
                if not tag:
                    continue

                try:
                    sentiment = tag.split("|")[3].strip().lower()
                except:
                    pass

                try:
                    category = tag.split("|")[2].strip().lower()
                except:
                    category = "NA"

                if sentiment == "positive":
                    positive += 1
                elif sentiment == "negative":
                    negative += 1
                if sentiment == "neutral":
                    neutral += 1

                comment_sentiment_dict.append(
                    {"Sentiment": sentiment, "Comment": tag, "Category": category}
                )
                split_tags_td.append(tag)

        comment_sentiment_df = pd.DataFrame(data=comment_sentiment_dict)
        for r in dataframe_to_rows(comment_sentiment_df, index=True, header=True):
            ws_comment_sentiment.append(r)

        ws_stats["A1"] = "Comments Processed"
        ws_stats["B1"] = self.all_comments_data["number_of_comments"]
        ws_stats["A2"] = "#Comments with replies"
        ws_stats["B2"] = self.all_comments_data["comments_with_replies"]
        ws_stats["A3"] = "%Comments with replies"
        ws_stats["B3"] = (
            self.all_comments_data["comments_with_replies"]
            / self.all_comments_data["number_of_comments"]
            * 100
        )
        ws_stats["A4"] = "Total replies"
        ws_stats["B4"] = self.all_comments_data["total_replies"]
        ws_stats["A5"] = "Average replies per comment"
        ws_stats["B5"] = self.all_comments_data["avg_replies"]
        ws_stats["A6"] = "Average likes per comment"
        ws_stats["B6"] = self.all_comments_data["avg_likes"]
        ws_stats["A7"] = "Average comment length"
        ws_stats["B7"] = self.all_comments_data["avg_comment_length"]

        ws_stats["A9"] = "% Positive"
        ws_stats["B9"] = str(
            float(positive) / self.all_comments_data["number_of_comments"] * 100
        )
        ws_stats["A10"] = "% Neutral"
        ws_stats["B10"] = str(
            float(neutral) / self.all_comments_data["number_of_comments"] * 100
        )
        ws_stats["A11"] = "% Negative"
        ws_stats["B11"] = str(
            float(negative) / self.all_comments_data["number_of_comments"] * 100
        )

        top_down_dict = dict(
            {k: [] for k in self.analysis_df["Top Down Topics"] if k != ""}
        )
        for comment in split_tags_td:
            flag = 0
            for topic in top_down_dict.keys():
                _topic = topic
                if _topic != None:
                    _topic = topic.lower()
                # topic_without_stopwords = ' '.join([word for word in topic.split() if word not in stopwords.words('english')])
                # if topic_without_stopwords.lower() in comment_topics.lower():
                if str(_topic) in comment.lower():
                    # if topic.lower()[:13] in comment.lower():
                    top_down_dict[topic].append(comment)
                    flag = 1
            if not flag:
                self.not_present["top_down"].append(comment)

        self.top_down_dict = top_down_dict
        top_down_df = []

        for k in top_down_dict.keys():
            for v in top_down_dict[k]:
                top_down_df.append({"Topic": k, "Comment": v})

        td_df = pd.DataFrame(data=top_down_df)
        # td_df.to_csv("Top_Down_Raw_" + self.videoID + ".csv")

        # top_down_stats = []
        # for k in top_down_dict.keys():
        #     top_down_stats.append({"Topic": k, "Comments": len(top_down_dict[k])})

        # td_stats = pd.DataFrame(data=top_down_stats)

        sentiment_dict = []
        for k, v in top_down_dict.items():
            # print ('Topic: ', k)
            mini_pos = 0
            mini_neg = 0
            mini_neu = 0
            if len(v) < 3:
                continue
            for comment in v:
                try:
                    sentiment = comment.split("|")[3].strip()
                except:
                    continue
                if sentiment.lower() == "positive":
                    mini_pos += 1
                elif sentiment.lower() == "negative":
                    mini_neg += 1
                elif sentiment.lower() == "neutral":
                    mini_neu += 1
                else:
                    print("None")
                    print(comment)
            if mini_pos > mini_neg and mini_pos > mini_neu:
                sentiment = "Positive"
            elif mini_neg > mini_pos and mini_neg > mini_neu:
                sentiment = "Negative"
            elif mini_neu > mini_pos and mini_neu > mini_neg:
                sentiment = "Neutral"
                # sentiment_dict.append({"Topic": k, "Sentiment": 'Neutral'})
            else:
                if mini_pos == mini_neg:
                    sentiment = "Positive, Negative"
                if mini_pos == mini_neu:
                    sentiment = "Positive, Neutral"
                if mini_neu == mini_neg:
                    sentiment = "Neutral, Negative"
                print("None largest")
            sentiment_dict.append(
                {"Topic": k, "Sentiment": sentiment, "Comments": len(v)}
            )

        topics_sentiment_df = pd.DataFrame(data=sentiment_dict)
        for r in dataframe_to_rows(topics_sentiment_df, index=True, header=True):
            ws_topics_match.append(r)

        virtual_workbook = NamedTemporaryFile(delete=True)
        workbook.save(virtual_workbook.name)
        return virtual_workbook

